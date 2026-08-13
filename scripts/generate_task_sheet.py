#!/usr/bin/env python3
"""Generate project task management spreadsheet for Google Sheets import."""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = Path(__file__).resolve().parents[1] / "docs" / "project-management" / "diem10-quan-ly-cong-viec.xlsx"

MEMBERS = ("Thành viên A", "Thành viên B", "Cả hai", "Chưa phân công")
STATUSES = ("Chưa làm", "Đang làm", "Chờ review", "Hoàn thành", "Tạm hoãn")
PRIORITIES = ("Cao", "Trung bình", "Thấp")
AREAS = ("Frontend", "Backend", "Full-stack", "DevOps", "Docs", "QA/UAT")

HEADER_FILL = PatternFill("solid", fgColor="1A73E8")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill("solid", fgColor="F8F9FA")

STATUS_FILLS = {
    "Chưa làm": PatternFill("solid", fgColor="E8EAED"),
    "Đang làm": PatternFill("solid", fgColor="FEF7E0"),
    "Chờ review": PatternFill("solid", fgColor="E8F0FE"),
    "Hoàn thành": PatternFill("solid", fgColor="E6F4EA"),
    "Tạm hoãn": PatternFill("solid", fgColor="FCE8E6"),
}

TASK_COLUMNS = [
    "ID",
    "Tên công việc",
    "Mô tả",
    "User Story",
    "Khu vực",
    "Người phụ trách",
    "Trạng thái",
    "Ưu tiên",
    "Deadline",
    "Ghi chú",
    "Cập nhật",
]

TASKS = [
    (
        "T-001",
        "Google OAuth login + session cookie",
        "US-02: PKCE, callback, tạo hồ sơ theo Google sub, logout",
        "US-02",
        "Full-stack",
        "Chưa phân công",
        "Chưa làm",
        "Cao",
        "",
        "Tiền đề cho mọi luồng học sinh",
        "",
    ),
    (
        "T-002",
        "Admin allowlist + RBAC + audit log",
        "US-03: email allowlist, chặn 403 non-admin, audit thay đổi quyền",
        "US-03",
        "Backend",
        "Chưa phân công",
        "Chưa làm",
        "Cao",
        "",
        "",
        "",
    ),
    (
        "T-003",
        "Danh sách & lọc đề đã xuất bản",
        "US-04: tìm kiếm, filter chủ đề/năm/mức độ, pagination",
        "US-04",
        "Full-stack",
        "Chưa phân công",
        "Chưa làm",
        "Cao",
        "",
        "",
        "",
    ),
    (
        "T-004",
        "Bắt đầu / tiếp tục lượt làm",
        "US-05: timer server-side, resume attempt",
        "US-05",
        "Full-stack",
        "Chưa phân công",
        "Hoàn thành",
        "Cao",
        "",
        "Đã có pause/resume server",
        "2026-08-08",
    ),
    (
        "T-005",
        "Autosave câu trả lời",
        "US-06: lưu tiến độ không mất khi refresh",
        "US-06",
        "Full-stack",
        "Chưa phân công",
        "Chưa làm",
        "Cao",
        "",
        "",
        "",
    ),
    (
        "T-006",
        "Hết giờ tự nộp bài",
        "US-07: auto-submit khi expires_at",
        "US-07",
        "Full-stack",
        "Chưa phân công",
        "Hoàn thành",
        "Cao",
        "",
        "Gắn với pause/resume",
        "2026-08-08",
    ),
    (
        "T-007",
        "Nộp bài + màn hình kết quả",
        "US-08: chấm điểm, hiển thị điểm và đáp án sau nộp",
        "US-08",
        "Full-stack",
        "Chưa phân công",
        "Chưa làm",
        "Cao",
        "",
        "",
        "",
    ),
    (
        "T-008",
        "Lịch sử làm bài & làm lại",
        "US-09: gom theo đề, xem snapshot, retake version published",
        "US-09",
        "Full-stack",
        "Chưa phân công",
        "Hoàn thành",
        "Trung bình",
        "",
        "21 test API đã pass",
        "2026-08-08",
    ),
    (
        "T-009",
        "Upload tài liệu nguồn (R2 presigned)",
        "US-10: DOCX/PDF upload, confirm checksum",
        "US-10",
        "Full-stack",
        "Chưa phân công",
        "Hoàn thành",
        "Trung bình",
        "",
        "",
        "2026-08-09",
    ),
    (
        "T-010",
        "Import parser → draft",
        "US-11: manual-exams DOCX, idempotency, reject OCR",
        "US-11",
        "Backend",
        "Chưa phân công",
        "Hoàn thành",
        "Trung bình",
        "",
        "30 test API",
        "2026-08-10",
    ),
    (
        "T-011",
        "Rà soát draft & publish",
        "US-12: editor, validation, publish version",
        "US-12",
        "Full-stack",
        "Chưa phân công",
        "Hoàn thành",
        "Trung bình",
        "",
        "40 test API",
        "2026-08-11",
    ),
    (
        "T-012",
        "Tạo draft thủ công (editor chung)",
        "US-13: draft trống hoặc từ đề published",
        "US-13",
        "Full-stack",
        "Chưa phân công",
        "Chưa làm",
        "Cao",
        "",
        "Backlog tiếp theo sau US-12",
        "",
    ),
    (
        "T-013",
        "Landing page & kho đề công khai",
        "US-01: metadata published, CTA, mobile",
        "US-01",
        "Frontend",
        "Chưa phân công",
        "Hoàn thành",
        "Thấp",
        "",
        "",
        "2026-07-27",
    ),
    (
        "T-014",
        "E2E Docker Compose — luồng học sinh",
        "Login → làm bài → nộp → lịch sử qua compose local",
        "Validation",
        "QA/UAT",
        "Chưa phân công",
        "Chưa làm",
        "Cao",
        "",
        "Theo initial-delivery plan",
        "",
    ),
    (
        "T-015",
        "E2E Docker Compose — luồng admin",
        "Upload → import → review → publish → draft thủ công",
        "Validation",
        "QA/UAT",
        "Chưa phân công",
        "Chưa làm",
        "Cao",
        "",
        "",
        "",
    ),
    (
        "T-016",
        "Chuẩn bị staging VPS",
        "R2, Nginx, env OAuth, migration trên Ubuntu 24.04",
        "Ops",
        "DevOps",
        "Chưa phân công",
        "Chưa làm",
        "Trung bình",
        "",
        "Decision 0001",
        "",
    ),
    (
        "T-017",
        "CI: lint + test + security scan",
        "Ruff, Pyright, Pytest, pnpm lint/build ổn định",
        "Ops",
        "DevOps",
        "Chưa phân công",
        "Đang làm",
        "Trung bình",
        "",
        "Scaffold đã có, duy trì xanh",
        "",
    ),
    (
        "T-018",
        "Cập nhật runbook vận hành",
        "docs/operations/runbook.md cho deploy & rollback",
        "Docs",
        "Docs",
        "Chưa phân công",
        "Chưa làm",
        "Thấp",
        "",
        "",
        "",
    ),
]


def style_header(ws, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def add_list_validation(ws, col_letter: str, options: tuple[str, ...], max_row: int = 200) -> None:
    formula = f'"{",".join(options)}"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Chọn giá trị trong danh sách"
    dv.errorTitle = "Giá trị không hợp lệ"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{max_row}")


def setup_tasks_sheet(ws) -> None:
    ws.title = "Công việc"
    ws.append(TASK_COLUMNS)
    for row in TASKS:
        ws.append(row)

    style_header(ws, len(TASK_COLUMNS))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(TASK_COLUMNS))}{len(TASKS) + 1}"

    widths = [8, 36, 42, 12, 14, 18, 14, 12, 12, 28, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    add_list_validation(ws, "E", AREAS)
    add_list_validation(ws, "F", MEMBERS)
    add_list_validation(ws, "G", STATUSES)
    add_list_validation(ws, "H", PRIORITIES)

    for row_idx in range(2, len(TASKS) + 2):
        if row_idx % 2 == 0:
            for col in range(1, len(TASK_COLUMNS) + 1):
                ws.cell(row=row_idx, column=col).fill = ALT_FILL
        for col in range(1, len(TASK_COLUMNS) + 1):
            ws.cell(row=row_idx, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    status_col = get_column_letter(TASK_COLUMNS.index("Trạng thái") + 1)
    for status, fill in STATUS_FILLS.items():
        ws.conditional_formatting.add(
            f"{status_col}2:{status_col}200",
            FormulaRule(formula=[f'${status_col}2="{status}"'], fill=fill),
        )


def setup_dashboard(ws, task_count: int) -> None:
    ws.title = "Tổng quan"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20

    title_font = Font(bold=True, size=16, color="1A73E8")
    ws["A1"] = "Diem10 Lịch Sử — Quản lý công việc"
    ws["A1"].font = title_font
    ws.merge_cells("A1:C1")

    rows = [
        ("", "", ""),
        ("Dự án", "Nền tảng học Lịch sử THPT (monorepo Next.js + FastAPI)", ""),
        ("Repo", "diem10lichsuv2", ""),
        ("Ngày tạo sheet", date.today().isoformat(), ""),
        ("Thành viên", "Đổi tên cột F: Thành viên A / Thành viên B", ""),
        ("", "", ""),
        ("Hướng dẫn nhanh", "", ""),
        ("1", "Tab Công việc: gán người, cập nhật trạng thái & deadline", ""),
        ("2", "Tab Tuần này: copy task ưu tiên tuần hiện tại", ""),
        ("3", "Tab Họp: ghi quyết định sau mỗi buổi sync", ""),
        ("4", "Upload file lên Google Drive → Mở bằng Google Sheets", ""),
        ("", "", ""),
        ("Thống kê (cập nhật thủ công hoặc dùng COUNTIF sau khi import)", "", ""),
        ("Tổng task", str(task_count), ""),
        ("Hoàn thành", "4 task (US-01,05,07,09,10,11,12 — kiểm tra lại tab Công việc)", ""),
        ("Còn lại", "US-02~04,06,08,13 + validation + staging", ""),
        ("Ưu tiên tuần", "Phân công US-02/03 hoặc US-13 tùy roadmap", ""),
    ]
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if c_idx == 1 and value and value[0].isdigit() is False and value not in ("", "Hướng dẫn nhanh", "Thống kê (cập nhật thủ công hoặc dùng COUNTIF sau khi import)"):
                cell.font = Font(bold=True)


def setup_weekly_sheet(ws) -> None:
    ws.title = "Tuần này"
    headers = ["Tuần", "ID tham chiếu", "Focus", "Người phụ trách", "Mục tiêu cuối tuần", "Trạng thái"]
    ws.append(headers)
    style_header(ws, len(headers))
    sample = [
        (f"W{date.today().isocalendar()[1]}-{date.today().year}", "T-012", "US-13 draft thủ công", "Thành viên A", "API + UI tạo draft trống", "Chưa làm"),
        (f"W{date.today().isocalendar()[1]}-{date.today().year}", "T-001", "Google OAuth", "Thành viên B", "Login/logout end-to-end", "Chưa làm"),
        (f"W{date.today().isocalendar()[1]}-{date.today().year}", "T-014", "E2E học sinh", "Cả hai", "Script hoặc checklist UAT", "Chưa làm"),
    ]
    for row in sample:
        ws.append(row)
    widths = [14, 14, 32, 18, 36, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    add_list_validation(ws, "D", MEMBERS)
    add_list_validation(ws, "F", STATUSES)


def setup_meeting_sheet(ws) -> None:
    ws.title = "Họp & quyết định"
    headers = ["Ngày", "Người tham dự", "Nội dung", "Quyết định", "Action items (ID task)", "Deadline"]
    ws.append(headers)
    style_header(ws, len(headers))
    ws.append([date.today().isoformat(), "A + B", "Khởi tạo sheet quản lý công việc", "Phân công US-13 vs auth theo ưu tiên", "T-012, T-001", ""])
    widths = [12, 18, 40, 40, 22, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def setup_us_backlog_sheet(ws) -> None:
    ws.title = "Backlog US"
    headers = ["User Story", "Tên", "Trạng thái dự án", "Ghi chú"]
    ws.append(headers)
    style_header(ws, len(headers))
    backlog = [
        ("US-01", "Khám phá sản phẩm", "Hoàn thành", "Landing + public exams API"),
        ("US-02", "Đăng nhập Google", "Chưa làm", ""),
        ("US-03", "Quyền admin", "Chưa làm", ""),
        ("US-04", "Tìm đề", "Chưa làm", ""),
        ("US-05", "Bắt đầu/tiếp tục attempt", "Hoàn thành", "Pause/resume server"),
        ("US-06", "Autosave", "Chưa làm", ""),
        ("US-07", "Hết giờ auto-submit", "Hoàn thành", ""),
        ("US-08", "Nộp & kết quả", "Chưa làm", ""),
        ("US-09", "Lịch sử & làm lại", "Hoàn thành", ""),
        ("US-10", "Upload nguồn", "Hoàn thành", ""),
        ("US-11", "Import draft", "Hoàn thành", ""),
        ("US-12", "Review & publish", "Hoàn thành", ""),
        ("US-13", "Draft thủ công", "Chưa làm", "Backlog tiếp theo"),
    ]
    for row in backlog:
        ws.append(row)
    for idx, width in enumerate([12, 28, 18, 40], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    setup_dashboard(wb.active, len(TASKS))
    setup_tasks_sheet(wb.create_sheet())
    setup_weekly_sheet(wb.create_sheet())
    setup_meeting_sheet(wb.create_sheet())
    setup_us_backlog_sheet(wb.create_sheet())
    wb.save(OUTPUT)
    print(f"Created: {OUTPUT}")


if __name__ == "__main__":
    main()
